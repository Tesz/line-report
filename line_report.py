#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LINE Report Generator
Generate report (markdown or excel) from LINE chat export files and images.

Usage: python line_report.py <Report Name> <Folder> <Sort> [Format]
  Sort: d (date) or n (name)
  Format: md (markdown) or xlsx (excel), default: md

Example: 
  python line_report.py 20260308 ./20260308 d        # output markdown
  python line_report.py 20260308 ./20260308 d xlsx   # output excel
"""

# === CONFIGURATION ===
# Media markers - supports multiple languages (array)
MEDIA_MARKERS = [
    "รูป",        # Thai
    "[Photo]",    # English
    "Photos",     # English
    "写真",       # Japanese
    "照片",       # Chinese
    "วิดีโอ",     # Thai video
    "Videos",      # English video
]

# Supported media extensions
IMAGE_EXTENSIONS = ['.jpg', '.jpeg', '.png', '.gif']
VIDEO_EXTENSIONS = ['.mp4', '.mov', '.avi']

# Job title prefix
JOB_TITLE = "งานที่"
# ====================

import os
import sys
import re
from pathlib import Path

# For Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Image sizing for Excel (10cm max width)
# Excel uses points as units: 1cm ≈ 28.35 points
MAX_IMAGE_WIDTH_CM = 10
MAX_IMAGE_WIDTH_PT = MAX_IMAGE_WIDTH_CM * 28.35  # ~283.5 points


def find_chatlog(folder):
    """Find chatlog file: [LINE]Keep Memo.txt or any single .txt file."""
    folder_path = Path(folder)
    
    # Try [LINE]Keep Memo.txt first (various patterns)
    patterns = [
        "[LINE]Keep Memo.txt",
        "[LINE] Chat with Keep Memo.txt",
        "[LINE] Chat history with Keep Memo.txt"
    ]
    
    for pattern in patterns:
        keep_memo = folder_path / pattern
        if keep_memo.exists():
            return keep_memo
    
    # Also try finding any file starting with [LINE]
    for f in folder_path.iterdir():
        if f.is_file() and f.suffix.lower() == '.txt' and f.name.startswith('[LINE]'):
            return f
    
    # Find all .txt files
    txt_files = list(folder_path.glob("*.txt"))
    
    if len(txt_files) == 0:
        print("Error: No .txt file found in folder")
        sys.exit(1)
    
    if len(txt_files) > 1:
        print("Error: Multiple .txt files found. Please use [LINE]Keep Memo.txt")
        sys.exit(1)
    
    return txt_files[0]


def get_media_by_name(folder):
    """Get all media files (images + videos) sorted by name (alphabetically, ascending)."""
    folder_path = Path(folder)
    media = []
    
    all_extensions = IMAGE_EXTENSIONS + VIDEO_EXTENSIONS
    
    for f in folder_path.iterdir():
        if f.is_file() and f.suffix.lower() in all_extensions:
            media.append((f.name, f.name))  # (sort_key, original_name)
    
    # Sort alphabetically by filename
    media.sort(key=lambda x: x[0])
    return media


def get_media_by_date(folder):
    """Get all media files sorted by creation date (oldest first)."""
    folder_path = Path(folder)
    media = []
    
    all_extensions = IMAGE_EXTENSIONS + VIDEO_EXTENSIONS
    
    for f in folder_path.iterdir():
        if f.is_file() and f.suffix.lower() in all_extensions:
            # Get creation time
            ctime = f.stat().st_ctime
            # Use original name as secondary sort for stability
            media.append((ctime, f.name))
    
    # Sort by date (ASC), then by name for stability
    media.sort(key=lambda x: (x[0], x[1]))
    return media


def extract_entries(chat_content):
    """Parse LINE chat log and extract entries with messages and media counts.
    
    Entries are grouped by content continuity:
    - If next message has NO text (only media), it will be grouped with previous entry
    
    Returns list of dict with keys: date, time, sender, message, media_count
    """
    entries = []
    
    lines = chat_content.split('\n')
    current_date = ""
    current_time = ""
    current_sender = ""
    current_message = ""
    media_count = 0
    
    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        
        # Check for date line: YYYY.MM.DD วัน[วัน]
        date_match = re.match(r'^(\d{4}\.\d{2}\.\d{2})\s+วัน', line_stripped)
        if date_match:
            current_date = date_match.group(1)
            continue
        
        # Check if this line starts with a timestamp (entry start)
        # Format: HH:MM Name Message or HH:MM Name รูป
        timestamp_match = re.match(r'^(\d{2}:\d{2})\s+(.+?)\s+"(.+)"', line_stripped)
        timestamp_no_quote_match = re.match(r'^(\d{2}:\d{2})\s+(.+?)\s+(.+)$', line_stripped)
        
        if timestamp_match:
            # Save previous entry if exists
            if current_message or media_count > 0:
                entries.append({
                    'date': current_date,
                    'time': current_time,
                    'sender': current_sender,
                    'message': current_message.strip(),
                    'media_count': media_count
                })
            
            # Start new entry with quoted message
            current_time = timestamp_match.group(1)
            current_sender = timestamp_match.group(2)
            current_message = timestamp_match.group(3)
            media_count = 0
            
        elif timestamp_no_quote_match:
            # Check if it's a media marker (photo/video)
            if timestamp_no_quote_match.group(3).strip() in MEDIA_MARKERS:
                # Media-only message: ADD to current entry (don't create new entry)
                if not current_time:  # First media has no time yet
                    current_time = timestamp_no_quote_match.group(1)
                    current_sender = timestamp_no_quote_match.group(2)
                media_count += 1
            else:
                # New message with text: save previous and start new
                if current_message or media_count > 0:
                    entries.append({
                        'date': current_date,
                        'time': current_time,
                        'sender': current_sender,
                        'message': current_message.strip(),
                        'media_count': media_count
                    })
                
                # Start new entry without quotes
                current_time = timestamp_no_quote_match.group(1)
                current_sender = timestamp_no_quote_match.group(2)
                current_message = timestamp_no_quote_match.group(3)
                media_count = 0
                
        elif line_stripped in MEDIA_MARKERS:
            # Media marker on separate line: ADD to current entry
            media_count += 1
            
        elif current_message:
            # Continuation of message (multi-line)
            if current_message and not current_message.endswith('\n'):
                current_message += '\n'
            current_message += line_stripped
    
    # Add last entry
    if current_message or media_count > 0:
        entries.append({
            'date': current_date,
            'time': current_time,
            'sender': current_sender,
            'message': current_message.strip(),
            'media_count': media_count
        })
    
    return entries


def interleave_entry_images(images):
    """Reorder images for a single entry: odd first, then even."""
    if len(images) <= 1:
        return images
    
    odd_images = []
    even_images = []
    
    for i, img in enumerate(images):
        if i % 2 == 0:  # 0, 2, 4 -> odd positions (1st, 3rd, 5th)
            odd_images.append(img)
        else:  # 1, 3, 5 -> even positions (2nd, 4th, 6th)
            even_images.append(img)
    
    # Odd then even
    return odd_images + even_images


def generate_markdown(entries, image_files, report_name, output_path):
    """Generate markdown report."""
    
    # Get image names in order (already sorted by name or date)
    # image_files is list of (sort_key, original_name)
    ordered_images = [img[1] for img in image_files]
    
    # Track image index
    img_idx = 0
    
    # Build markdown content
    md_content = f"# {report_name}\n\n"
    
    for i, entry in enumerate(entries, 1):
        md_content += f"## {JOB_TITLE} {i}\n"
        md_content += f"{entry['message']}\n\n"
        
        media_count = entry['media_count']
        
        if media_count > 0:
            # Get images for this entry
            entry_images = []
            for _ in range(media_count):
                if img_idx < len(ordered_images):
                    entry_images.append(ordered_images[img_idx])
                    img_idx += 1
                else:
                    entry_images.append("[ไม่พบรูป]")
            
            # Interleave within this entry
            entry_images = interleave_entry_images(entry_images)
            
            # Split into odd and even columns
            half = (len(entry_images) + 1) // 2
            odd_imgs = entry_images[:half]
            even_imgs = entry_images[half:]
            
            # Build markdown with proper col blocks
            md_content += "````col\n"
            
            # Column 1: odd images
            md_content += "```col-md\n"
            for img in odd_imgs:
                md_content += f"![[{img}]]\n"
            md_content += "```\n"
            
            # Column 2: even images
            if even_imgs:
                md_content += "```col-md\n"
                for img in even_imgs:
                    md_content += f"![[{img}]]\n"
                md_content += "```\n"
            
            md_content += "````\n\n"
    
    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(md_content)
    
    print(f"✓ Report generated: {output_path}")
    print(f"  - Entries: {len(entries)}")
    print(f"  - Images used: {img_idx}")


def generate_excel(entries, image_files, report_name, output_path, folder):
    """Generate Excel report with Cover and Detail sheets."""
    
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is not installed. Please install with: pip install openpyxl")
        sys.exit(1)
    
    # Get image files sorted (name or date)
    ordered_images = [img[1] for img in image_files]
    
    # Track image index
    img_idx = 0
    
    # Get all image paths for the folder
    folder_path = Path(folder)
    
    # Create workbook
    wb = Workbook()
    
    # ============================================
    # Sheet 1: "Cover" - Data table
    # ============================================
    ws_cover = wb.active
    ws_cover.title = "Cover"
    
    # Header row
    headers = ["No.", "Date", "Time", "Sender", "Text Content", "Image Count", "Image Filenames", "Image Path"]
    for col, header in enumerate(headers, 1):
        cell = ws_cover.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    # Add data rows
    for row, entry in enumerate(entries, 2):
        # Get images for this entry
        entry_images = []
        for _ in range(entry['media_count']):
            if img_idx < len(ordered_images):
                entry_images.append(ordered_images[img_idx])
                img_idx += 1
        
        # Write row data
        ws_cover.cell(row=row, column=1, value=row-1)  # No.
        ws_cover.cell(row=row, column=2, value=entry.get('date', ''))  # Date
        ws_cover.cell(row=row, column=3, value=entry.get('time', ''))  # Time
        ws_cover.cell(row=row, column=4, value=entry.get('sender', ''))  # Sender
        ws_cover.cell(row=row, column=5, value=entry.get('message', ''))  # Text Content
        ws_cover.cell(row=row, column=6, value=entry['media_count'])  # Image Count
        ws_cover.cell(row=row, column=7, value=", ".join(entry_images) if entry_images else "")  # Image Filenames
        ws_cover.cell(row=row, column=8, value=f"{folder}/")  # Image Path
        
        # Enable text wrapping
        for col in range(1, 9):
            ws_cover.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    
    # Auto-adjust column width for Cover sheet
    for col in range(1, 9):
        max_length = 0
        column_letter = ws_cover.cell(row=1, column=col).column_letter
        for cell in ws_cover[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_cover.column_dimensions[column_letter].width = adjusted_width
    
    # ============================================
    # Sheet 2: "Detail" - Text + embedded images
    # ============================================
    ws_detail = wb.create_sheet(title="Detail")
    
    # Reset image index for Detail sheet
    img_idx = 0
    current_row = 1
    
    for entry_idx, entry in enumerate(entries, 1):
        # Get images for this entry
        entry_images = []
        for _ in range(entry['media_count']):
            if img_idx < len(ordered_images):
                entry_images.append(ordered_images[img_idx])
                img_idx += 1
        
        # === Entry Header ===
        header_text = f"งานที่ {entry_idx}"
        if entry.get('date'):
            header_text += f" - {entry['date']}"
        if entry.get('time'):
            header_text += f" {entry['time']}"
        if entry.get('sender'):
            header_text += f" - {entry['sender']}"
        
        cell_header = ws_detail.cell(row=current_row, column=1, value=header_text)
        cell_header.font = Font(bold=True, size=12)
        current_row += 1
        
        # === Text Content ===
        if entry.get('message'):
            cell_text = ws_detail.cell(row=current_row, column=1, value=entry['message'])
            cell_text.font = Font(size=10)
            cell_text.alignment = Alignment(wrap_text=True, vertical='top')
            ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            current_row += 1
        
        # === Images (stacked vertically) ===
        for img_name in entry_images:
            img_path = folder_path / img_name
            if img_path.exists():
                try:
                    # Load and add image
                    img = XLImage(str(img_path))
                    
                    # Calculate dimensions to fit max width
                    original_width = img.width
                    original_height = img.height
                    
                    # Scale if width exceeds max
                    if original_width > MAX_IMAGE_WIDTH_PT:
                        scale = MAX_IMAGE_WIDTH_PT / original_width
                        img.width = MAX_IMAGE_WIDTH_PT
                        img.height = original_height * scale
                    
                    # Add image to cell
                    cell = ws_detail.cell(row=current_row, column=1)
                    ws_detail.row_dimensions[current_row].height = img.height * 0.75
                    img.anchor = f'A{current_row}'
                    ws_detail.add_image(img)
                    current_row += 1
                except Exception as e:
                    # If image fails to load, skip
                    print(f"  Warning: Could not load image {img_name}: {e}")
                    current_row += 1
            else:
                # Image not found, add placeholder text
                cell = ws_detail.cell(row=current_row, column=1, value=f"[ไม่พบรูป: {img_name}]")
                current_row += 1
        
        # === Spacing row between entries ===
        current_row += 1
    
    # Set column width for Detail sheet
    ws_detail.column_dimensions['A'].width = 40
    ws_detail.column_dimensions['B'].width = 40
    
    # Save file
    wb.save(output_path)
    
    print(f"✓ Excel report generated: {output_path}")
    print(f"  - Entries: {len(entries)}")
    print(f"  - Images used: {img_idx}")
    print(f"  - Sheets: Cover, Detail")


def main():
    if len(sys.argv) < 4 or len(sys.argv) > 5:
        print("Usage: python line_report.py <Report Name> <Folder> <Sort> [Format]")
        print("  Sort: d (date) or n (name)")
        print("  Format: md (markdown) or xlsx (excel), default: md")
        print("Example:")
        print("  python line_report.py 20260308 ./20260308 d        # output markdown")
        print("  python line_report.py 20260308 ./20260308 d xlsx   # output excel")
        sys.exit(1)
    
    report_name = sys.argv[1]
    folder = sys.argv[2]
    sort_option = sys.argv[3].lower()
    output_format = sys.argv[4].lower() if len(sys.argv) == 5 else 'md'
    
    # Validate sort option
    if sort_option not in ['d', 'n']:
        print("Error: Sort must be 'd' (date) or 'n' (name)")
        sys.exit(1)
    
    # Validate format option
    if output_format not in ['md', 'xlsx']:
        print("Error: Format must be 'md' (markdown) or 'xlsx' (excel)")
        sys.exit(1)
    
    # Check xlsx dependency
    if output_format == 'xlsx' and not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is not installed. Please install with: pip install openpyxl")
        sys.exit(1)
    
    # Find chatlog
    print(f"Finding chatlog in: {folder}")
    chatlog_path = find_chatlog(folder)
    print(f"Using chatlog: {chatlog_path.name}")
    
    # Read chatlog
    with open(chatlog_path, 'r', encoding='utf-8') as f:
        chat_content = f.read()
    
    # Parse entries
    entries = extract_entries(chat_content)
    print(f"Found {len(entries)} entries")
    
    # Get images based on sort option
    if sort_option == 'd':
        image_files = get_media_by_date(folder)
        print(f"Found {len(image_files)} images (sorted by date)")
    else:
        image_files = get_media_by_name(folder)
        print(f"Found {len(image_files)} images (sorted by name)")
    
    # Generate output based on format
    if output_format == 'xlsx':
        output_path = Path(folder) / f"{report_name}.xlsx"
        generate_excel(entries, image_files, report_name, output_path, folder)
    else:
        output_path = Path(folder) / f"{report_name}.md"
        generate_markdown(entries, image_files, report_name, output_path)


if __name__ == "__main__":
    main()